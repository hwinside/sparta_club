from openpyxl import load_workbook

work_book = load_workbook('prac01.xlsx')
work_sheet = work_book['prac']

# 데이터를 읽어봅니다.
# temp_text = work_sheet.cell(row = 1, column = 3).value

work_sheet.cell(row=2, column=2, value='홍길동')

# 수정한 엑셀파일을 저장합니다.
# 참고: 다른이름으로 저장할 수도 있습니다.
work_book.save('prac01.xlsx')